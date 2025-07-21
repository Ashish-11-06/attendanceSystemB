import io
import mimetypes
import os
import random
import sys
import tempfile
from tkinter import Image
# from turtle import pd
import pandas as pd 
import uuid
from wsgiref.util import FileWrapper 
from datetime import datetime
import zipfile
from django.http import FileResponse, StreamingHttpResponse
import openpyxl
from pdf2image import convert_from_path
from rest_framework.parsers import MultiPartParser, FormParser
from django.utils.dateparse import parse_date
from django.db.models import Q
import pytesseract
import imghdr
import cv2
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.hashers import make_password, check_password

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from management.models import Admin, Attendance, AttendanceFile, EventUnitLocation, Events, Khetra, Location, Register, Unit, Volunteer
from management.serializers import AdminSerializer, AttendanceFileSerializer, AttendanceFileUploadSerializer, AttendanceSerializer, EventUnitLocationSerializer, EventsSerializer, KhetraSerializer, LocationSerializer, LoginSerializer, RegisterSerializer, UnitSerializer, VolunteerSerializer
from management.utils import extract_table_data_from_image, parse_sewadal_adhikari_data,  send_otp_email
from rest_framework_simplejwt.tokens import AccessToken
from rest_framework.permissions import IsAuthenticated
from rest_framework.permissions import AllowAny

# class RegisterAPIView(APIView):
#     def post(self, request):
#         serializer = UnitSerializer(data=request.data)
#         if serializer.is_valid():
#             # Generate OTP
#             otp = str(random.randint(100000, 999999))

#             # Temporarily save OTP in validated_data
#             serializer.save(otp=otp)

#             # Send OTP to email
#             email = serializer.validated_data.get('email')
#             send_otp_email(email, otp)

#             return Response({
#                 "message": "Registration successful! OTP sent to email.",
#                 "data": serializer.data
#             }, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RegisterAPIView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        data = request.data.copy()

        # Hash the password before saving
        if 'password' in data:
            raw_password = data['password']
            data['plain_password'] = raw_password  # Store raw if needed
            data['password'] = make_password(raw_password)

        serializer = UnitSerializer(data=data)
        if serializer.is_valid():
            otp = str(random.randint(100000, 999999))
            serializer.save(otp=otp)

            email = serializer.validated_data.get('email')
            send_otp_email(email, otp)

            return Response({
                "message": "Registration successful! OTP sent to email.",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class VerifyOTPAPIView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        email = request.data.get("email")
        otp = request.data.get("otp")

        if not email or not otp:
            return Response({"error": "Email and OTP are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = Unit.objects.get(email=email, otp=otp)
            # OTP matched — now clear it from the database
            user.otp = None  # or use '' if you prefer empty string
            user.save()
            
            return Response({"message": "OTP verification successful!"}, status=status.HTTP_200_OK)
        except Unit.DoesNotExist:
            return Response({"error": "Invalid email or OTP."}, status=status.HTTP_400_BAD_REQUEST)


def get_token_for_user(user_instance, user_type):
    token = AccessToken.for_user(user_instance)
    token['user_type'] = user_type  # add custom claim
    return token

        
class LoginAPIView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        email = serializer.validated_data['email'].strip().lower()
        password = serializer.validated_data['password'].strip()

        user_type = serializer.validated_data['user_type']

        if user_type == 'admin':
            try:
                admin = Admin.objects.get(email__iexact=email)
                if check_password(password, admin.password):
                    access_token = get_token_for_user(admin, 'admin')
                    return Response({
                        "message": "Admin login successful",
                        "user": {
                            "id": admin.id,
                            "admin_name": admin.name,
                            "email": admin.email,
                            "access": str(access_token),
                        }
                    }, status=status.HTTP_200_OK)
                else:
                    return Response({"error": "Invalid admin credentials"}, status=status.HTTP_401_UNAUTHORIZED)
            except Admin.DoesNotExist:
                return Response({"error": "Invalid admin credentials"}, status=status.HTTP_401_UNAUTHORIZED)

        elif user_type == 'unit':
            # Try Register table first
            user = Register.objects.filter(email__iexact=email, user_type='unit').first()
            if user and check_password(password, user.password):
                access_token = get_token_for_user(user, 'unit')
                return Response({
                    "message": "Unit login successful ",
                    "user": {
                        "id": user.id,
                        "unit_name": user.unit_name,
                        "email": user.email,
                        "user_type": user.user_type,
                        "access": str(access_token),
                        
                    }
                }, status=status.HTTP_200_OK)

            # Try Unit table next
            unit = Unit.objects.filter(email__iexact=email).first()
            if unit and check_password(password, unit.password):
                access_token = get_token_for_user(unit, 'unit')
                return Response({
                    "message": "Unit login successful",
                    "user": {
                        "id": unit.id,
                        "unit_name": unit.unit_name,
                        "email": unit.email,
                        "user_type": "unit",
                        "access": str(access_token),
                    }
                }, status=status.HTTP_200_OK)

            # If both fail
            return Response({"error": "Invalid unit credentials"}, status=status.HTTP_401_UNAUTHORIZED)

        return Response({"error": "Invalid user_type"}, status=status.HTTP_400_BAD_REQUEST) 
    
  
  
    
    # Ashish --
# class UploadVolunteerExcelView(APIView):
#     def post(self, request, unit_id):
#         file = request.FILES.get('file')
#         if not file:
#             return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             unit_from_url = Unit.objects.get(id=unit_id)
#         except Unit.DoesNotExist:
#             return Response({'error': f"Unit ID {unit_id} not found"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             wb = openpyxl.load_workbook(file)
#             inserted = {}

#             for sheet in wb.worksheets:
#                 sheet_name = sheet.title.strip().lower()
#                 count = 0
#                 updated = 0

#                 gender = None
#                 is_registered = True

#                 if sheet_name == 'gents':
#                     gender = 'Male'
#                 elif sheet_name == 'ladies':
#                     gender = 'Female'
#                 elif sheet_name == 'un-reg gents':
#                     gender = 'Male'
#                     is_registered = False
#                 elif sheet_name == 'un-reg ladies':
#                     gender = 'Female'
#                     is_registered = False

#                 raw_headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[3]]

#                 column_map = {
#                     'new p#': 'new_personal_number',
#                     'old p#': 'old_personal_number',
#                     'name': 'name',
#                     'contact no.': 'phone',
#                     'email': 'email',
#                     'volunteer id': 'volunteer_id',
#                     'unit id': 'unit_id',
#                     's#': 's_number'
#                 }

#                 header_field_map = {
#                     idx: column_map.get(header)
#                     for idx, header in enumerate(raw_headers)
#                     if column_map.get(header)
#                 }

#                 for row in sheet.iter_rows(min_row=4, values_only=True):
#                     if not any(row):
#                         continue

#                     row_data = {
#                         header_field_map[idx]: value
#                         for idx, value in enumerate(row)
#                         if header_field_map.get(idx)
#                     }

#                     # if row_data.get('s_number') == 'X':
#                     #     continue
                    
#                     is_active = row_data.get('s_number') != 'X'

#                     name = row_data.get('name', '').strip()
#                     if not name:
#                         continue

#                     new_pn = row_data.get('new_personal_number')
#                     new_phone = row_data.get('phone', '')

#                     if not new_pn and not new_phone:
#                         continue  # Skip if both identifiers are missing

#                     # Determine unit instance
#                     unit = unit_from_url
#                     unit_id_in_row = row_data.get('unit_id')
#                     if unit_id_in_row:
#                         try:
#                             unit = Unit.objects.get(id=unit_id_in_row)
#                         except Unit.DoesNotExist:
#                             print(f"Unit ID {unit_id_in_row} not found, using default unit")

#                     # Try to find existing volunteer
#                     volunteer = None
#                     if new_pn:
#                         volunteer = Volunteer.objects.filter(unit=unit, new_personal_number=new_pn).first()
#                     else:
#                         volunteer = Volunteer.objects.filter(unit=unit, phone=new_phone, name=name).first()

#                     if volunteer:
#                         # Update existing volunteer
#                         volunteer.name = name
#                         volunteer.email = row_data.get('email')
#                         volunteer.phone = new_phone
#                         volunteer.old_personal_number = row_data.get('old_personal_number')
#                         volunteer.gender = gender
#                         volunteer.is_registered = is_registered
#                         volunteer.is_active = is_active
#                         volunteer.save()
#                         updated += 1
#                     else:
#                         # Create new volunteer
#                         data = {
#                             'name': name,
#                             'email': row_data.get('email'),
#                             'phone': new_phone,
#                             'old_personal_number': row_data.get('old_personal_number'),
#                             'new_personal_number': new_pn,
#                             'gender': gender,
#                             'unit': unit.id if unit else None,
#                             'is_registered': is_registered,
#                             'is_active': is_active,
#                         }

#                         serializer = VolunteerSerializer(data=data)
#                         if serializer.is_valid():
#                             serializer.save()
#                             count += 1
#                         else:
#                             # print(f"Validation errors: {serializer.errors}")
#                             print(f"Row skipped due to validation errors: {serializer.errors}, data: {data}")


#                 inserted[sheet.title] = f"{count} inserted, {updated} updated"

#             return Response({'message': 'Volinteers data imported successfully!', 'details': inserted}, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             import traceback
#             traceback.print_exc()
#             raise e


class UploadVolunteerExcelView(APIView):
    def post(self, request, unit_id):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            unit_from_url = Unit.objects.get(id=unit_id)
        except Unit.DoesNotExist:
            return Response({'error': f"Unit ID {unit_id} not found"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            inserted = {}
            inserted_names = set()  # ✅ Track inserted names globally

            for sheet in wb.worksheets:
                sheet_name = sheet.title.strip().lower()
                
                 # ✅ Skip Adhikaris sheet completely
                if sheet_name == 'adhikaris':
                    inserted[sheet.title] = "Skipped (Adhikaris sheet)"
                    continue
                
                count = 0
                updated = 0

                gender = None
                is_registered = True

                if sheet_name == 'gents':
                    gender = 'Male'
                elif sheet_name == 'ladies':
                    gender = 'Female'
                elif sheet_name == 'un-reg gents':
                    gender = 'Male'
                    is_registered = False
                elif sheet_name == 'un-reg ladies':
                    gender = 'Female'
                    is_registered = False

                raw_headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[3]]

                column_map = {
                    'new p#': 'new_personal_number',
                    'old p#': 'old_personal_number',
                    'name': 'name',
                    'contact no.': 'phone',
                    'email': 'email',
                    'volunteer id': 'volunteer_id',
                    'unit id': 'unit_id',
                    's#': 's_number'
                }

                header_field_map = {
                    idx: column_map.get(header)
                    for idx, header in enumerate(raw_headers)
                    if column_map.get(header)
                }

                for row in sheet.iter_rows(min_row=4, values_only=True):
                    if not any(row):
                        continue

                    row_data = {
                        header_field_map[idx]: value
                        for idx, value in enumerate(row)
                        if header_field_map.get(idx)
                    }

                    is_active = row_data.get('s_number') != 'X'
                    name = row_data.get('name', '').strip()
                    if not name:
                        continue

                    # ✅ Skip duplicates in Adhikaris sheet
                    if sheet.title.strip().lower() == 'adhikaris' and name.lower() in inserted_names:
                        continue  # Skip this row

                    new_pn = row_data.get('new_personal_number')
                    new_phone = row_data.get('phone', '')

                    if not new_pn and not new_phone:
                        continue  # Skip if both identifiers are missing

                    unit = unit_from_url
                    unit_id_in_row = row_data.get('unit_id')
                    if unit_id_in_row:
                        try:
                            unit = Unit.objects.get(id=unit_id_in_row)
                        except Unit.DoesNotExist:
                            print(f"Unit ID {unit_id_in_row} not found, using default unit")

                    volunteer = None
                    if new_pn:
                        volunteer = Volunteer.objects.filter(unit=unit, new_personal_number=new_pn).first()
                    else:
                        volunteer = Volunteer.objects.filter(unit=unit, phone=new_phone, name=name).first()

                    if volunteer:
                        # Update existing volunteer
                        volunteer.name = name
                        volunteer.email = row_data.get('email')
                        volunteer.phone = new_phone
                        volunteer.old_personal_number = row_data.get('old_personal_number')
                        volunteer.gender = gender
                        volunteer.is_registered = is_registered
                        volunteer.is_active = is_active
                        volunteer.save()
                        updated += 1
                        inserted_names.add(name.lower())  # ✅ track updated names too
                    else:
                        # Create new volunteer
                        data = {
                            'name': name,
                            'email': row_data.get('email'),
                            'phone': new_phone,
                            'old_personal_number': row_data.get('old_personal_number'),
                            'new_personal_number': new_pn,
                            'gender': gender,
                            'unit': unit.id if unit else None,
                            'is_registered': is_registered,
                            'is_active': is_active,
                        }

                        serializer = VolunteerSerializer(data=data)
                        if serializer.is_valid():
                            serializer.save()
                            count += 1
                            inserted_names.add(name.lower())  # ✅ track inserted names
                        else:
                            print(f"Row skipped due to validation errors: {serializer.errors}, data: {data}")


                inserted[sheet.title] = f"{count} inserted, {updated} updated"

            return Response({'message': 'Volunteer data imported successfully!', 'details': inserted}, status=status.HTTP_201_CREATED)

        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e




class EventsAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, event_id=None):
        if event_id:
            events = Events.objects.filter(event_id=event_id)
        else:
            events = Events.objects.all().order_by('-created_at')

        result = []

        for event in events:
            
            event_data = {
                "id": event.id,
                "event_id": event.event_id,
                "event_name": event.event_name,
                "start_date": event.start_date,
                "end_date": event.end_date,
                "time": event.time,
                "locations": []
            }

            
            print("event data", event_data)
            event_locations = EventUnitLocation.objects.filter(event=event)
            location_ids = set(event_locations.values_list("location_id", flat=True))
            location_map = {}

            print("event locaions",event_locations)
            for location_id in location_ids:
            
                try:
                    location = Location.objects.get(id=location_id)
                    location_map[location_id] = {
                        "location_id": location.id,
                        "location_city": location.city,
                        "location_address": location.address,  
                        "units": []
                    }
                    print("location_map",location_map)
                except Location.DoesNotExist:
                    print(False)
                    continue

            for entry in event_locations:
                if entry.unit and entry.location_id in location_map:
                    unit = entry.unit
                    unit_data = {
                        "id": unit.id,
                        "unit_id": unit.unit_id,
                        "unit_name": unit.unit_name,
                        "email": unit.email,
                        "phone": unit.phone,
                        "location": unit.location
                    }
                    location_map[entry.location_id]["units"].append(unit_data)

            event_data["locations"] = list(location_map.values())
            result.append(event_data)

        return Response(result, status=200)



    def post(self, request):
            data = request.data.copy()

            # Generate unique event_id
            event_id = str(uuid.uuid4())[:8]
            data['event_id'] = event_id

            # Extract locations with units
            locations_data = data.pop("locations", [])

            # Serialize and save event
            serializer = EventsSerializer(data=data)
            if serializer.is_valid():
                event = serializer.save()

                # Save EventUnitLocation entries
                for loc in locations_data:
                    location = None
                    unit_ids = loc.get("unit", [])

                    # Case 1: Existing location by ID
                    if "location_id" in loc:
                        try:
                            location = Location.objects.get(id=loc["location_id"])
                        except Location.DoesNotExist:
                            return Response({"error": f"Location ID {loc['location_id']} not found."}, status=400)

                    # Case 2: New location to be created
                    elif "location" in loc:
                        location_data = loc["location"]
                        unique_location_id = str(uuid.uuid4())[:8]
                        location = Location.objects.create(
                            location_id=unique_location_id,
                            state=location_data.get("state", ""),
                            city=location_data.get("city", ""),
                            address=location_data.get("address", "")
                        )
                    else:
                        return Response({"error": "Each location must include either 'location_id' or 'location'."}, status=400)

                    # Link units or save location only (unit_ids may be empty)
                    if unit_ids:
                        for unit_id in unit_ids:
                            try:
                                unit = Unit.objects.get(id=unit_id)
                                EventUnitLocation.objects.create(
                                    event=event,
                                    location=location,
                                    unit=unit
                                )
                            except Unit.DoesNotExist:
                                return Response({"error": f"Unit ID {unit_id} not found."}, status=400)
                    else:
                        # Save location with no units
                        EventUnitLocation.objects.create(
                            event=event,
                            location=location,
                            unit=None  # Unit is optional
                        )

                # Prepare response
                response_data = {
                    "id": event.id,
                    "event_id": event.event_id,
                    "event_name": event.event_name,
                    "start_date": event.start_date,
                    "end_date": event.end_date,
                    "time": event.time,
                    "locations": []
                }

                # Group locations with units
                event_locations = EventUnitLocation.objects.filter(event=event)
                location_map = {}

                for entry in event_locations:
                    if not entry.location:
                        continue

                    loc_id = entry.location.id
                    if loc_id not in location_map:
                        location_map[loc_id] = {
                            "location_id": loc_id,
                            "location_city": entry.location.city,
                            "location_address": entry.location.address,
                            "units": []
                        }

                    if entry.unit:
                        location_map[loc_id]["units"].append({
                            "unit_id": entry.unit.id,
                            "unit_name": entry.unit.unit_name,
                            "email": entry.unit.email,
                            "phone": entry.unit.phone,
                            "location": entry.unit.location
                        })

                response_data["locations"] = list(location_map.values())

                return Response({
                    "message": "Event created successfully!",
                    "data": response_data
                }, status=201)

            return Response(serializer.errors, status=400)
        
    def put(self, request, event_id=None):
        if not event_id:
            return Response({"error": "event_id is required in URL."}, status=status.HTTP_400_BAD_REQUEST)

        event = get_object_or_404(Events, event_id=event_id)
        data = request.data.copy()
        locations_data = data.pop("locations", [])

        serializer = EventsSerializer(event, data=data)
        if serializer.is_valid():
            event = serializer.save()

            # Clear existing mappings
            EventUnitLocation.objects.filter(event=event).delete()

            location_map = {}

            for loc in locations_data:
                location_id = loc.get("location_id")
                unit_ids = loc.get("units") or loc.get("unit") or []

                try:
                    location = Location.objects.get(id=location_id)

                    # Initialize location in map
                    location_map[location_id] = {
                        "location_id": location.id,
                        "location_city": location.city,
                        "location_address": location.address,
                        "units": []
                    }

                    if unit_ids:
                        for unit_id in unit_ids:
                            try:
                                unit = Unit.objects.get(id=unit_id)
                                EventUnitLocation.objects.create(event=event, location=location, unit=unit)

                                location_map[location_id]["units"].append({
                                    "id": unit.id,
                                    "unit_id": unit.unit_id,
                                    "unit_name": unit.unit_name,
                                    "email": unit.email,
                                    "phone": unit.phone,
                                    "location": unit.location
                                })

                            except Unit.DoesNotExist:
                                return Response({"error": f"Unit ID {unit_id} not found."}, status=400)
                    else:
                        # ✅ Save location mapping even if no unit
                        EventUnitLocation.objects.create(event=event, location=location, unit=None)

                except Location.DoesNotExist:
                    return Response({"error": f"Location ID {location_id} not found."}, status=400)
                except Exception as e:
                    return Response({"error": f"Error while saving location/unit: {str(e)}"}, status=500)

            response_data = {
                "id": event.id,
                "event_id": event.event_id,
                "event_name": event.event_name,
                "start_date": event.start_date,
                "end_date": event.end_date,
                "time": event.time,
                "locations": list(location_map.values())
            }

            return Response({
                "message": "Event updated successfully!",
                "data": response_data
            }, status=200)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    def delete(self, request, event_id=None):
        if not event_id:
            return Response({"error": "event_id is required in URL."}, status=status.HTTP_400_BAD_REQUEST)
        event = get_object_or_404(Events, event_id=event_id)
        event.delete()
        return Response({"message": "Event deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
  
    
    
class LocationAPIView(APIView):
    def get(self, request, location_id=None):
        if location_id:
            # Get single event by id
            location = get_object_or_404(Location, location_id=location_id)
            serializer = LocationSerializer(location)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            # Get all events
            locations = Location.objects.all()
            serializer = LocationSerializer(locations, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = LocationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()  # location_id auto-generated in serializer.create()
            return Response({"message": "Location created successfully!", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, location_id=None):
        if not location_id:
            return Response({"error": "ID is required for update."}, status=status.HTTP_400_BAD_REQUEST)

        location = get_object_or_404(Location, location_id=location_id)
        serializer = LocationSerializer(location, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    
class UnitAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, unit_id=None):
        if unit_id:
            # Get single unit by id
            unit = get_object_or_404(Unit, unit_id=unit_id)
            serializer = UnitSerializer(unit)
            return Response(serializer.data, status=status.HTTP_200_OK)
        # Get all units
        else:
            units = Unit.objects.all()
            serializer = UnitSerializer(units, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        data = request.data.copy()
        if 'password' in data:
            raw_password = data['password']
            data['plain_password'] = raw_password
            data['password'] = make_password(raw_password)
        serializer = UnitSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Unit created successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, unit_id):
        if not unit_id:
            return Response({"error": "Unit ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        unit = get_object_or_404(Unit, id=unit_id)
        
        data = request.data.copy()  # Make a mutable copy of request data

        # If 'password' is present and not empty, hash and store it
        if data.get('password'):
            raw_password = data['password']
            data['plain_password'] = raw_password
            data['password'] = make_password(raw_password)
        else:
            # If password not provided or is empty, exclude it from update
            data.pop('password', None)

        serializer = UnitSerializer(unit, data=data, partial=True)  # partial=True allows skipping fields
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    

    def delete(self, request, unit_id=None):
        if not unit_id:
            return Response({"error": "unit_id is required in URL."}, status=status.HTTP_400_BAD_REQUEST)
        unit = get_object_or_404(Unit, unit_id=unit_id)
        unit.delete()
        return Response({"message": "Unit deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    

class VolunteerAPIView(APIView):
    def get(self, request, volunteer_id=None):
        if volunteer_id:
            volunteer = get_object_or_404(Volunteer, volunteer_id=volunteer_id)
            serializer = VolunteerSerializer(volunteer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        # Get all volunteers
        else:
            volunteers = Volunteer.objects.filter(is_active=True)
            serializer = VolunteerSerializer(volunteers, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = VolunteerSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Volunteer created successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, volunteer_id=None):
        if not volunteer_id:
            return Response({"error": "ID is required for update."}, status=status.HTTP_400_BAD_REQUEST)

        volunteer = get_object_or_404(Volunteer, volunteer_id=volunteer_id)
        serializer = VolunteerSerializer(volunteer, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    
class AdminAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, admin_id=None):
        if admin_id:
            # Get single admin by id
            volunteer = get_object_or_404(Admin, admin_id=admin_id)
            serializer = AdminSerializer(volunteer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        # Get all admins
        else:
            admins = Admin.objects.all()
            serializer = AdminSerializer(admins, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = AdminSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Admin created successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, admin_id=None):
        if not admin_id:
            return Response({"error": "ID is required for update."}, status=status.HTTP_400_BAD_REQUEST)

        admin = get_object_or_404(Admin, id=admin_id)
        data = request.data.copy()

        # If 'password' is not provided or is empty, remove it so it's not updated
        if not data.get('password'):
            data.pop('password', None)

        serializer = AdminSerializer(admin, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AttendanceAPIView(APIView):
    def get(self, request, attendance_id=None):
        if attendance_id:
            # Get single attendance by id
            attendance = get_object_or_404(Attendance, atd_id=attendance_id)
            serializer = AttendanceSerializer(attendance)
            return Response(serializer.data, status=status.HTTP_200_OK)
        # Get all attendance
        else:
            attendance = Attendance.objects.all()
            serializer = AttendanceSerializer(attendance, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

    # def post(self, request):
    #     serializer = AttendanceSerializer(data=request.data)
        
    #     if serializer.is_valid():
    #         serializer.save()
    #         return Response({"message": "Attendance recorded", "data": serializer.data}, status=status.HTTP_201_CREATED)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def post(self, request):
        is_many = isinstance(request.data, list)
        data = request.data if is_many else [request.data]  # Ensure a list for iteration

        # Validate duplicates before saving
        for item in data:
            event_id = item.get('event')
            unit_id = item.get('unit')
            date = item.get('date')  # Make sure 'date' is passed in correct format (YYYY-MM-DD)

            if Attendance.objects.filter(event_id=event_id, unit_id=unit_id, date=date).exists():
                return Response({
                    "message": f"Attendance already submitted for event ID {event_id}, unit ID {unit_id} on {date}."
                }, status=status.HTTP_400_BAD_REQUEST)

        # Proceed to save since no duplicates found
        serializer = AttendanceSerializer(data=request.data, many=is_many)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Attendance recorded successfully.",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "message": "Failed to record attendance.",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request):
        is_many = isinstance(request.data, list)
        data = request.data if is_many else [request.data]

        updated_items = []
        errors = []

        for item in data:
            event_id = item.get('event')
            unit_id = item.get('unit')
            volunteer_id = item.get('volunteer')  # This is internal DB ID (int)

            if not event_id or not unit_id or not volunteer_id:
                errors.append({
                    "message": "Missing 'event', 'unit', or 'volunteer' (ID) in one of the items",
                    "data": item
                })
                continue

            attendance = Attendance.objects.filter(
                event_id=event_id,
                unit_id=unit_id,
                volunteer_id=volunteer_id  # direct match with FK ID
            ).first()

            if not attendance:
                errors.append({
                    "message": f"Attendance record not found for event {event_id}, unit {unit_id}, and volunteer ID {volunteer_id}",
                    "data": item
                })
                continue

            serializer = AttendanceSerializer(attendance, data=item)
            if serializer.is_valid():
                serializer.save()
                updated_items.append(serializer.data)
            else:
                errors.append({
                    "message": f"Validation failed for volunteer ID {volunteer_id}",
                    "errors": serializer.errors
                })

        if errors and updated_items:
            return Response({
                "message": "Partial success: some attendance records updated, some failed.",
                "updated": updated_items,
                "errors": errors
            }, status=status.HTTP_207_MULTI_STATUS)

        elif errors:
            return Response({
                "message": "Failed to update attendance records.",
                "errors": errors
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "message": "All attendance records updated successfully.",
            "data": updated_items
        }, status=status.HTTP_200_OK)



class AttendanceFileAPIView(APIView):
    def get(self, request):
        event_id = request.query_params.get('event')
        unit_id = request.query_params.get('unit')
        
        files = AttendanceFile.objects.all()

        if event_id:
            files = files.filter(event_id=event_id)  # assuming event is a ForeignKey named event
        if unit_id:
            files = files.filter(unit_id=unit_id)    # assuming unit is a ForeignKey named unit

        serializer = AttendanceFileSerializer(files, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

    def post(self, request):
        # print('hello')
        serializer = AttendanceFileSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "File uploaded successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
 
 
    
class TotalCountAPIView(APIView):
    def get(self, request):
        total_events = Events.objects.count()
        total_volunteers = Volunteer.objects.filter( is_active=True).count()
        total_units = Unit.objects.count()
        total_locations = Location.objects.count()

        return Response({
            "total_events": total_events,
            "total_volunteers": total_volunteers,
            "total_units": total_units,
            "total_locations": total_locations
        }, status=status.HTTP_200_OK)

# class VolunteerCountAPIVIew(APIView):
#     def get(self, request):
#         total = Volunteer.objects.count()
#         return Response({"total_volunteers": total}, status=status.HTTP_200_OK)
    
# class UnitCountAPIView(APIView):
#     def get(self, request):
#         total = Unit.objects.count()
#         return Response({"total_units": total}, status=status.HTTP_200_OK)
    
# class LocationCountAPIView(APIView):
#     def get(self, request):
#         total = Location.objects.count()
#         return Response({"total_locations": total}, status=status.HTTP_200_OK)  
    
    
class DataFechEvenUnitIdAPIView(APIView):
    # def post (self, request, unit=None, event=None):
    #     unit = request.data.get('unit')
    #     event = request.data.get('event')
        
    #     if unit and event:
    #         # Get attendance for specific unit and event
    #         attendance = Attendance.objects.filter(unit_id=unit, event_id=event)
    #         serializer = AttendanceSerializer(attendance, many=True)
    #         return Response(serializer.data, status=status.HTTP_200_OK)
    #     else:
    #         return Response({"error": "unit_id and event_id are required."}, status=status.HTTP_400_BAD_REQUEST) 
        
    # def put(self, request):
    #     data = request.data
    #     if not isinstance(data, list) or not data:
    #         return Response({
    #             "message": "Expected a non-empty list of attendance records."
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     # Extract event and unit from first item
    #     event_id = data[0].get('event')
    #     unit_id = data[0].get('unit')

    #     if not event_id or not unit_id:
    #         return Response({
    #             "message": "Missing 'event' or 'unit' in the data."
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     # Delete existing records for this event and unit
    #     Attendance.objects.filter(event_id=event_id, unit_id=unit_id).delete()

    #     created_items = []
    #     errors = []

    #     for index, item in enumerate(data):
    #         serializer = AttendanceSerializer(data=item)
    #         if serializer.is_valid():
    #             serializer.save()
    #             created_items.append(serializer.data)
    #         else:
    #             errors.append({
    #                 "index": index,
    #                 "errors": serializer.errors,
    #                 "data": item
    #             })

    #     if errors and created_items:
    #         return Response({
    #             "message": "Partial success: some records created, others failed.",
    #             "created": created_items,
    #             "errors": errors
    #         }, status=status.HTTP_207_MULTI_STATUS)

    #     elif errors:
    #         return Response({
    #             "message": "Failed to create any records.",
    #             "errors": errors
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     return Response({
    #         "message": "All attendance records updated successfully.",
    #         "data": created_items
    #     }, status=status.HTTP_200_OK)


    def post(self, request):
        unit = request.data.get('unit')
        event = request.data.get('event')
        date = request.data.get('date')  # Expecting a list of date strings

        if not unit or not event or not date:
            return Response({
                "error": "unit, event, and date (list) are required."
            }, status=status.HTTP_400_BAD_REQUEST)

        if not isinstance(date, list):
            return Response({
                "error": "date must be a list of 'YYYY-MM-DD' strings."
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Parse date strings to date objects
            date_objs = [datetime.strptime(d, '%Y-%m-%d').date() for d in date]
        except ValueError:
            return Response({
                "error": "One or more dates are in invalid format. Use 'YYYY-MM-DD'."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Fetch only updated attendance records for given event, unit, and dates
        attendance = Attendance.objects.filter(
            unit_id=unit,
            event_id=event,
            date__in=date_objs,
            already_updated=True
        ).order_by('date', 'volunteer_id')  # Optional sorting

        serializer = AttendanceSerializer(attendance, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    
    def put(self, request):
        data = request.data
        if not isinstance(data, list) or not data:
            return Response({
                "message": "Expected a non-empty list of attendance records."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Extract event and unit from first item
        event_id = data[0].get('event')
        unit_id = data[0].get('unit')

        if not event_id or not unit_id:
            return Response({
                "message": "Missing 'event' or 'unit' in the data."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Optional: Clear existing attendance for those dates and volunteers
        date = [item.get("date") for item in data if item.get("date")]
        Attendance.objects.filter(event_id=event_id, unit_id=unit_id, date__in=date).delete()

        created_items = []
        errors = []

        for index, item in enumerate(data):
            serializer = AttendanceSerializer(data=item)
            if serializer.is_valid():
                serializer.save()
                created_items.append(serializer.data)
            else:
                errors.append({
                    "index": index,
                    "errors": serializer.errors,
                    "data": item
                })

        if errors and created_items:
            return Response({
                "message": "Partial success: some records created, others failed.",
                "created": created_items,
                "errors": errors
            }, status=status.HTTP_207_MULTI_STATUS)

        elif errors:
            return Response({
                "message": "Failed to create any records.",
                "errors": errors
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "message": "All attendance records saved successfully.",
            "data": created_items
        }, status=status.HTTP_200_OK) 
    
    
    # def post(self, request, unit=None, event=None):
    #     unit = request.data.get('unit')
    #     event = request.data.get('event')
    #     start_date = request.data.get('start_date')
    #     end_date = request.data.get('end_date')

    #     # ✅ CASE 1: Date Range Filtering
    #     if unit and start_date and end_date:
    #         try:
    #             start = parse_date(start_date)
    #             end = parse_date(end_date)

    #             if not start or not end:
    #                 return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

    #             # ✅ Get all events for this unit that overlap the date range
    #             event_ids = Events.objects.filter(
    #                 Q(start_date__lte=end),
    #                 Q(end_date__gte=start),
    #                 units__id=unit
    #             ).values_list('id', flat=True)

    #             if not event_ids:
    #                 return Response({
    #                     "message": "No events found for this unit in the given date range."
    #                 }, status=status.HTTP_404_NOT_FOUND)

    #             # ✅ Get attendance for all matching events
    #             attendance = Attendance.objects.filter(
    #                 unit_id=unit,
    #                 event_id__in=event_ids
    #             )

    #             serializer = AttendanceSerializer(attendance, many=True)

    #             return Response({
    #                 "unit_id": unit,
    #                 "start_date": start_date,
    #                 "end_date": end_date,
    #                 "total_events": len(event_ids),
    #                 "attendance_records": serializer.data
    #             }, status=status.HTTP_200_OK)

    #         except Exception as e:
    #             return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    #     # ✅ CASE 2: Specific Unit + Event
    #     elif unit and event:
    #         attendance = Attendance.objects.filter(unit_id=unit, event_id=event)
    #         serializer = AttendanceSerializer(attendance, many=True)
    #         return Response(serializer.data, status=status.HTTP_200_OK)

    #     return Response(
    #         {"error": "Please provide 'unit' and either 'event' or 'start_date' + 'end_date'."},
    #         status=status.HTTP_400_BAD_REQUEST
    #     )
        
    # def put(self, request):
    #     data = request.data
    #     if not isinstance(data, list) or not data:
    #         return Response({
    #             "message": "Expected a non-empty list of attendance records."
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     # 🔹 Optional: support date range filter for events
    #     start_date = request.query_params.get('start_date')
    #     end_date = request.query_params.get('end_date')

    #     try:
    #         if start_date and end_date:
    #             start = parse_date(start_date)
    #             end = parse_date(end_date)
    #             if not start or not end:
    #                 raise ValueError("Invalid date format.")

    #     except Exception as e:
    #         return Response({"error": f"Invalid date range: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    #     created_items = []
    #     errors = []

    #     for index, item in enumerate(data):
    #         event_id = item.get('event')
    #         unit_id = item.get('unit')

    #         if not event_id or not unit_id:
    #             errors.append({
    #                 "index": index,
    #                 "error": "Missing 'event' or 'unit' in the record.",
    #                 "data": item
    #             })
    #             continue

    #         # ✅ Optional: if date range provided, check if event is within it
    #         if start_date and end_date:
    #             try:
    #                 event = Events.objects.get(id=event_id)
    #                 if event.start_date > end or (event.end_date and event.end_date < start):
    #                     errors.append({
    #                         "index": index,
    #                         "error": f"Event {event_id} not in date range.",
    #                         "data": item
    #                     })
    #                     continue
    #             except Events.DoesNotExist:
    #                 errors.append({
    #                     "index": index,
    #                     "error": f"Event {event_id} does not exist.",
    #                     "data": item
    #                 })
    #                 continue

    #         # 🔹 Delete previous attendance for this event and unit (only once)
    #         if not any(att.event_id == event_id and att.unit_id == unit_id for att in created_items):
    #             Attendance.objects.filter(event_id=event_id, unit_id=unit_id).delete()

    #         serializer = AttendanceSerializer(data=item)
    #         if serializer.is_valid():
    #             instance = serializer.save()
    #             created_items.append(instance)
    #         else:
    #             errors.append({
    #                 "index": index,
    #                 "errors": serializer.errors,
    #                 "data": item
    #             })

    #     if errors and created_items:
    #         return Response({
    #             "message": "Partial success: some records created, others failed.",
    #             "created": [AttendanceSerializer(obj).data for obj in created_items],
    #             "errors": errors
    #         }, status=status.HTTP_207_MULTI_STATUS)

    #     elif errors:
    #         return Response({
    #             "message": "Failed to create any records.",
    #             "errors": errors
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     return Response({
    #         "message": "All attendance records saved successfully.",
    #         "created": [AttendanceSerializer(obj).data for obj in created_items]
    #     }, status=status.HTTP_200_OK)
        
        
class EventUnitLocationAPIView(APIView):
    def get(self, request):
        locations = EventUnitLocation.objects.all()
        serializer = EventUnitLocationSerializer(locations, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # def post(self, request):
    #     serializer = EventUnitLocationSerializer(data=request.data)
    #     if serializer.is_valid():
    #         serializer.save()
    #         return Response({"message": "Event-Unit-Location created successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class AttendanceFileDownloadAPIView(APIView):
    def post(self, request):
        try:
            event_id = request.data.get('event')
            unit_id = request.data.get('unit')

            if not event_id or not unit_id:
                return Response({"error": "Missing event/unit ID"}, status=status.HTTP_400_BAD_REQUEST)

            attendance_file = AttendanceFile.objects.get(event__id=event_id, unit__id=unit_id)
            file_path = attendance_file.file.path

            if not os.path.exists(file_path):
                return Response({"error": "File not found on server"}, status=status.HTTP_404_NOT_FOUND)

            filename = os.path.basename(file_path)
            file_ext = os.path.splitext(filename)[1].lower()

            content_types = {
                '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                '.xls': 'application/vnd.ms-excel',
                '.pdf': 'application/pdf',
                '.csv': 'text/csv',
                '.txt': 'text/plain',
                '.doc': 'application/msword',
                '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.png': 'image/png',
            }

            content_type = content_types.get(file_ext, 'application/octet-stream')

            file = open(file_path, 'rb')
            response = FileResponse(file, content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Cache-Control'] = 'no-store'

            return response

        except AttendanceFile.DoesNotExist:
            return Response({"error": "File record not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 
        
            
class VolunteersByUnitPostAPIView(APIView):
   
    
    def post(self, request):
        unit_id = request.data.get("unit")

        if not unit_id:
            return Response({"error": "Unit ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        volunteers = Volunteer.objects.filter(unit_id=unit_id, is_active=True).order_by('name')

        if not volunteers.exists():
            return Response({"message": "No volunteers found for this unit."}, status=status.HTTP_404_NOT_FOUND)

        result = []

        for volunteer in volunteers:
            # Fetch all attendance records for this volunteer under the same unit
            attendance_records = Attendance.objects.filter(volunteer=volunteer, unit_id=unit_id).order_by('-date')

            attendance_data = [
                {
                    "event_id": att.id,
                    "date": att.date,
                    "present": att.present,
                    "absent": att.absent,
                    "in_time": att.in_time,
                    "out_time": att.out_time,
                    "remark": att.remark
                }
                for att in attendance_records
            ]

            result.append({
                "id": volunteer.id,
                "name": volunteer.name,
                "new_personal_number": volunteer.new_personal_number,
                "gender": volunteer.gender,  
                'is_registered': volunteer.is_registered,
                # "mobile": volunteer.mobile,
                "attendance": attendance_data
            })

        return Response(result, status=status.HTTP_200_OK)
    

# class UploadVolunteerImageStatsView(APIView):
#     def post(self, request):
#         upload = request.FILES.get('file')
#         if not upload:
#             return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

#         ext = upload.name.split('.')[-1].lower()
#         if ext not in ['pdf', 'png', 'jpg', 'jpeg', 'bmp']:
#             return Response({'error': 'Invalid file format. Only PDF, PNG, JPG, JPEG, BMP allowed.'},
#                             status=status.HTTP_400_BAD_REQUEST)

#         try:
#             # Save uploaded file temporarily
#             with tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}") as temp_file:
#                 for chunk in upload.chunks():
#                     temp_file.write(chunk)
#                 temp_path = temp_file.name

#             all_stats = []

#             if ext == 'pdf':
#                 # Convert PDF pages to images
#                 images = convert_from_path(temp_path, poppler_path=r'D:\\Prushal\\poppler-24.08.0\\Library\\bin')

#                 for img in images:  
#                     with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as img_file:
#                         img.save(img_file.name, 'PNG')
#                         text = extract_table_data_from_image(img_file.name)
#                         stats = parse_ocr_output_to_counts(text)
#                         all_stats.extend(stats)
#                         os.remove(img_file.name)

#             else:
#                 # For image files
#                 image = cv2.imread(temp_path)
#                 if image is None:
#                     os.remove(temp_path)
#                     return Response({'error': 'OpenCV failed to read the image.'}, status=status.HTTP_400_BAD_REQUEST)
#                 text = extract_table_data_from_image(temp_path)
#                 all_stats = parse_ocr_output_to_counts(text)

#             os.remove(temp_path)
#             return Response({'data': all_stats}, status=status.HTTP_200_OK)

#         except Exception as e:
#             if 'temp_path' in locals() and os.path.exists(temp_path):
#                 os.remove(temp_path)
#             return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class UploadFileExtractTextAPIView(APIView):
    
    
    def post(self, request):
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            return Response({'error': 'No file uploaded.'}, status=400)

        try:
            ext = uploaded_file.name.split('.')[-1].lower()

            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}") as temp_file:
                for chunk in uploaded_file.chunks():
                    temp_file.write(chunk)
                temp_path = temp_file.name

            all_text = []

            if ext == "pdf":
                images = convert_from_path(temp_path, poppler_path=r"D:\Prushal\poppler-24.08.0\Library\bin")
                for image in images:
                    text = pytesseract.image_to_string(image, config="--psm 6")
                    lines = [line.strip() for line in text.split('\n') if line.strip()]
                    all_text.extend(lines)

            elif ext in ["png", "jpg", "jpeg"]:
                image = Image.open(temp_path)
                text = pytesseract.image_to_string(image, config="--psm 6")
                lines = [line.strip() for line in text.split('\n') if line.strip()]
                all_text.extend(lines)

            elif ext in ["xlsx", "xls"]:
                df = pd.read_excel(temp_path)
                for _, row in df.iterrows():
                    row_values = [str(val) for val in row.tolist()]
                    all_text.append(" ".join(row_values))

            else:
                return Response({"error": "Unsupported file format."}, status=400)

            os.remove(temp_path)

            # Parse lines into structured table only for PDF/Image OCR text
            if ext in ["pdf", "png", "jpg", "jpeg"]:
                structured_data = parse_sewadal_adhikari_data(all_text)


                return Response({"data": structured_data}, status=200)

            # For Excel: show extracted rows as-is
            return Response({"data": all_text}, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=500)
        
        
class   VolunteersReportAPIView(APIView):
    def get(self, request):
        try:
            data = []
            units = Unit.objects.all()

            for unit in units:
                volunteers = Volunteer.objects.filter(unit=unit)

                # Counts for all volunteers (active + inactive)
                total_male = volunteers.filter(gender='Male').count()
                total_female = volunteers.filter(gender='Female').count()
                # total_registered = volunteers.filter(is_registered=True).count()
                active_gents = volunteers.filter(is_active=True, gender='Male').count()
                active_ladies = volunteers.filter(is_active=True, gender='Female').count()
                total_registered = total_male + total_female
                
                active_total = active_gents + active_ladies

                total_unregistered = volunteers.filter(is_registered=False).count()
                unregistered_male = volunteers.filter(is_registered=False, gender='Male').count()
                unregistered_female = volunteers.filter(is_registered=False, gender='Female').count()
                # grand_total = volunteers.count()  
                
                grand_total = total_registered + total_unregistered

                # Counts for active volunteers only
                # total_active = volunteers.filter(is_active=True).count()
                # print((total_active))
                # Print names of all active volunteers for testing
                # active_volunteers = volunteers.filter(is_active=True)
                # print("Active volunteer names:", [v.name for v in active_volunteers])

                # data.append({
                #     "id": unit.id,
                #     "khetra": unit.khetra.khetra if unit.khetra else None,
                #     "unit_id": unit.unit_id,
                #     "unit_name": unit.unit_name,

                #     "total_male": total_male,
                #     "total_female": total_female,
                #     "total_registered": total_registered,
                #     "total_unregistered": total_unregistered,
                #     "unregistered_male": unregistered_male,
                #     "unregistered_female": unregistered_female,

                #     "total_active": total_active,
                #     "grand_total": grand_total
                # })
                data.append({
                    "id": unit.id,
                    "khetra": unit.khetra.khetra if unit.khetra else None,
                    "unit_id": unit.unit_id,
                    "unit_name": unit.unit_name,

                    "reg_gents": total_male,
                    "reg_ladies": total_female,
                    "active_gents": active_gents,
                    "active_ladies": active_ladies,
                    "active_total": active_total,
                    "unreg_gents": unregistered_male,
                    "unreg_ladies": unregistered_female,

                    "grand_total": grand_total
                })

            return Response(data, status=200)

        except Exception as e:
            return Response({"status": False, "message": str(e)}, status=500)
     
     
     
from collections import defaultdict

class AttendanceReportAPIView(APIView):
    def get(self, request, event_id=None):
        try:
            if event_id is None:
                return Response({"message": "Event ID is required"}, status=400)

            # Get event info
            event = get_object_or_404(Events, id=event_id)
            event_name = event.event_name

            # Get all present records for the event
            attendance_qs = Attendance.objects.filter(event_id=event_id, present=True)

            # Get distinct dates from attendance
            distinct_dates = attendance_qs.values_list('date', flat=True).distinct()
            response_data = []

            for date in distinct_dates:
                date_wise_units = attendance_qs.filter(date=date).values_list('unit_id', flat=True).distinct()
                
                for unit_id in date_wise_units:
                    unit_attendance = attendance_qs.filter(date=date, unit_id=unit_id)

                    total_present = unit_attendance.count()
                    total_present_male = unit_attendance.filter(volunteer__gender__iexact="Male").count()
                    total_present_female = unit_attendance.filter(volunteer__gender__iexact="Female").count()

                    total_unregister_male = unit_attendance.filter(
                        volunteer__gender__iexact="Male",
                        volunteer__is_registered=False
                    ).count()

                    total_unregister_female = unit_attendance.filter(
                        volunteer__gender__iexact="Female",
                        volunteer__is_registered=False
                    ).count()

                    total_reg_male_present = total_present_male - total_unregister_male
                    total_reg_female_present = total_present_female - total_unregister_female

                    total_reg = total_reg_male_present + total_reg_female_present
                    total_unreg = total_unregister_male + total_unregister_female
                    grand_total = total_reg + total_unreg

                    unit_name = Unit.objects.get(id=unit_id).unit_name

                    response_data.append({
                        "date": str(date),
                        "unit_id": unit_id,
                        "unit_name": unit_name,
                        "total_present": total_present,
                        "total_present_male": total_present_male,
                        "total_present_female": total_present_female,
                        "total_unregister_male": total_unregister_male,
                        "total_unregister_female": total_unregister_female,
                        "total_register_male": total_reg_male_present,
                        "total_register_female": total_reg_female_present,
                        "total_register": total_reg,
                        "total_unregister": total_unreg,
                        "grand_total": grand_total
                    })

            return Response({
                "event_id": event_id,
                "event_name": event_name,
                "datewise_unit_summary": response_data
            }, status=200)

        except Exception as e:
            return Response({"status": False, "message": str(e)}, status=500)

class KhetraAPIView(APIView):

    def get(self, request):
    
        khetras = Khetra.objects.all()
        serializer = KhetraSerializer(khetras, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):

        serializer = KhetraSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Khetra added successfully.", "data": serializer.data}, status=201)
        return Response(serializer.errors, status=200)
    
    
class OverallVolunteersStatsAPIView(APIView):
    def get(self, request):
        try:
            volunteers = Volunteer.objects.all()

            total_male = volunteers.filter(gender='Male').count()
            total_female = volunteers.filter(gender='Female').count()
            total_registered = volunteers.filter(is_registered=True).count()
            total_unregistered = volunteers.filter(is_registered=False).count()
            unregistered_male = volunteers.filter(is_registered=False, gender='Male').count()
            unregistered_female = volunteers.filter(is_registered=False, gender='Female').count()
            # total_active = volunteers.filter(is_active=True).count()
            # total_inactive = volunteers.filter(is_active=False).count()
            # grand_total = volunteers.count()

            return Response({
                "total_male": total_male,
                "total_female": total_female,
                "total_registered": total_registered,
                "total_unregistered": total_unregistered,
                "unregistered_male": unregistered_male,
                "unregistered_female": unregistered_female,
                # "total_active": total_active,
                # "total_inactive": total_inactive,
                # "grand_total": grand_total
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"status": False, "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        
class VolunteersPendingAttendanceAPIView(APIView):
     def post(self, request):
        unit_id = request.data.get("unit")
        event_id = request.data.get("event")
        date = request.data.get("date")

        if not (unit_id and event_id and date):
            return Response({
                "error": "unit, event, and date are required."
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            attendance_date = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return Response({
                "error": "Invalid date format. Use YYYY-MM-DD."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Get all attendance records for given event, unit, and date
        attendance_qs = Attendance.objects.filter(
            unit_id=unit_id,
            event_id=event_id,
            date=attendance_date
        ).select_related('volunteer')

        if not attendance_qs.exists():
            return Response({
                "message": "No volunteers have been marked for attendance yet."
            }, status=status.HTTP_200_OK)

        serializer = AttendanceSerializer(attendance_qs, many=True)
        return Response({
            "message": "Volunteers already marked for attendance.",
            "data": serializer.data
        }, status=status.HTTP_200_OK)
        
        
class AttendanceFileUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        serializer = AttendanceFileUploadSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Attendance file uploaded successfully",
                "data": serializer.data
            }, status=201)
        return Response(serializer.errors, status=400)

    def get(self, request):
        event_id = request.query_params.get('event')
        unit_id = request.query_params.get('unit')
        
        queryset = AttendanceFile.objects.all()
        if event_id:
            queryset = queryset.filter(event_id=event_id)
        if unit_id:
            queryset = queryset.filter(unit_id=unit_id)

        serializer = AttendanceFileUploadSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)
